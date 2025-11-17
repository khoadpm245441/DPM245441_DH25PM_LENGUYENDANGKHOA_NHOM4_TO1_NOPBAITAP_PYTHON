import openpyxl
import os

FILE_PATH = "NhanVien.xlsx"
HEADER = ["STT", "Mã", "Tên", "Tuổi"]
SHEET_NAME = "NhanVien"

# =================================================================
# I. CÁC HÀM XỬ LÝ EXCEL
# =================================================================

def load_or_create_workbook(file_path, sheet_name, header):
    """Tải workbook hiện có hoặc tạo mới nếu file chưa tồn tại."""
    if os.path.exists(file_path):
        # Tải file hiện có
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(header)
        else:
            worksheet = workbook[sheet_name]
    else:
        # Tạo file mới và ghi tiêu đề
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet.append(header)
    return workbook, worksheet

def read_data_from_excel(worksheet):
    """Đọc tất cả nhân viên từ worksheet và trả về dạng list of dicts."""
    employees = []
    # Bỏ qua hàng tiêu đề (hàng 1)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
        if all(cell is None for cell in row): # Bỏ qua dòng trống
            continue

        try:
            # Dữ liệu: STT, Mã, Tên, Tuổi (chỉ số 0, 1, 2, 3)
            employees.append({
                'STT': int(row[0]) if row[0] is not None else row_index + 1,
                'Ma': str(row[1]) if row[1] is not None else '',
                'Ten': str(row[2]) if row[2] is not None else '',
                'Tuoi': int(row[3]) if row[3] is not None else 0
            })
        except (ValueError, TypeError):
            # Xử lý lỗi nếu dữ liệu không đúng định dạng (ví dụ: tuổi không phải số)
            continue
    return employees

# =================================================================
# II. CHỨC NĂNG CHÍNH
# =================================================================

# --- 1. Lưu (Ghi nối) Nhân viên vào File Excel ---
def luu_nhan_vien():
    """Nhập thông tin NV và ghi nối vào file Excel."""
    
    workbook, worksheet = load_or_create_workbook(FILE_PATH, SHEET_NAME, HEADER)
    
    print("\n--- NHẬP THÔNG TIN NHÂN VIÊN ---")
    ma = input("Nhập Mã NV (ví dụ: NV7): ").upper()
    ten = input("Nhập Tên NV: ")
    while True:
        try:
            tuoi = int(input("Nhập Tuổi: "))
            if tuoi > 0:
                break
            else:
                print("Tuổi phải lớn hơn 0.")
        except ValueError:
            print("Lỗi: Tuổi phải là số nguyên.")

    # STT là số hàng tiếp theo
    stt = worksheet.max_row
    if stt == 1 and all(cell.value is None for cell in worksheet[1]): # Nếu sheet rỗng hoàn toàn
         stt = 1
    elif worksheet.max_row > 0:
         # Vì hàng 1 là header, hàng tiếp theo là max_row + 1. Nhưng do ta dùng header từ đầu nên max_row đã tính cả header
         stt = worksheet.max_row 
         # Nếu file có dữ liệu, stt mới sẽ là hàng tiếp theo, tức là bằng số hàng hiện có
         # Cần lấy STT lớn nhất của cột A và cộng 1. Ở đây đơn giản là dùng số hàng hiện tại.
         stt = stt 
    
    # Dữ liệu mới: STT, Mã, Tên, Tuổi
    new_data = [stt, ma, ten, tuoi]
    
    # Ghi dữ liệu vào hàng cuối cùng
    worksheet.append(new_data)
    
    # Lưu file
    try:
        workbook.save(FILE_PATH)
        print(f"✅ Đã lưu nhân viên '{ten}' vào file '{FILE_PATH}'.")
    except Exception as e:
        print(f"❌ Lỗi khi lưu file: {e}. Vui lòng đóng file Excel nếu đang mở.")


# --- 2. Đọc và Xuất danh sách Nhân viên ---
def doc_va_xuat_danh_sach(employees_list=None):
    """Đọc dữ liệu từ file và in ra màn hình."""
    
    workbook, worksheet = load_or_create_workbook(FILE_PATH, SHEET_NAME, HEADER)
    
    # Nếu không có list đầu vào (đang đọc toàn bộ), thì đọc từ Excel
    if employees_list is None:
        employees = read_data_from_excel(worksheet)
        title = "DANH SÁCH NHÂN VIÊN TỪ FILE EXCEL"
    else:
        employees = employees_list
        title = "KẾT QUẢ SẮP XẾP/TÌM KIẾM"

    print("\n" + "=" * 40)
    print(title)
    print("-" * 40)
    print(f"| {'STT':<5} | {'Mã':<5} | {'Tên':<15} | {'Tuổi':<5} |")
    print("-" * 40)

    if not employees:
        print("| Danh sách trống.                       |")
    else:
        for emp in employees:
            print(f"| {emp['STT']:<5} | {emp['Ma']:<5} | {emp['Ten']:<15} | {emp['Tuoi']:<5} |")
            
    print("=" * 40)


# --- 3. Sắp xếp Nhân viên theo Tuổi tăng dần ---
def sap_xep_nhan_vien():
    """Sắp xếp nhân viên theo tuổi tăng dần."""
    
    workbook, worksheet = load_or_create_workbook(FILE_PATH, SHEET_NAME, HEADER)
    employees = read_data_from_excel(worksheet)
    
    if not employees:
        print("❌ Danh sách nhân viên trống, không thể sắp xếp.")
        return

    # Sắp xếp theo 'Tuổi' tăng dần (mặc định là False cho tăng dần)
    employees_sorted = sorted(employees, key=lambda x: x['Tuoi'], reverse=False)
    
    print("\n--- NHÂN VIÊN SAU KHI SẮP XẾP THEO TUỔI TĂNG DẦN ---")
    doc_va_xuat_danh_sach(employees_sorted)


# =================================================================
# III. MENU CHÍNH
# =================================================================

def menu_chinh():
    """Hiển thị menu và xử lý lựa chọn của người dùng."""
    while True:
        print("\n" + "#" * 40)
        print("PHẦN MỀM QUẢN LÝ NHÂN VIÊN (EXCEL FILE)")
        print("#" * 40)
        print("1. Lưu (Thêm mới) Nhân viên vào File")
        print("2. Đọc (Xuất danh sách) Nhân viên")
        print("3. Sắp xếp Nhân viên theo Tuổi tăng dần")
        print("0. Thoát chương trình")
        print("-" * 40)
        
        choice = input("Vui lòng chọn chức năng: ")
        
        if choice == '1':
            luu_nhan_vien()
        elif choice == '2':
            doc_va_xuat_danh_sach()
        elif choice == '3':
            sap_xep_nhan_vien()
        elif choice == '0':
            print("Chương trình kết thúc. Tạm biệt!")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng chọn lại.")

# Khởi động chương trình
if __name__ == "__main__":
    menu_chinh()